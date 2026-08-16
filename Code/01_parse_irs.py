#Imports
import re
import zipfile
import pandas as pd
import xlrd
from openpyxl import load_workbook
from pathlib import Path

# IRS county-to-county migration for Hawaii, 1990-91 through 2022-23
# The IRS changed the file format 3 times over 33 years (holy headache) so
# there is one cleaning function per format, then everything gets assembled
# into one long dataframe at the bottom, same as my ACS cleanup did.
#
# year convention: year = FIRST year of the pair, so the 1990-1991 file is
# year 1990 (people who moved between filing 1990 and filing 1991)
#
# Kalawao County (FIPS 005, the tiny Kalaupapa peninsula, ~80 people) gets
# added into Maui County so the counties match the DBEDT tourism data.
# Its numbers are suppressed (-1) most years anyway.

#make sure the output folder exists
Path("data/processed").mkdir(parents=True, exist_ok=True)

#where the data lives
RAW_ZIP = "data/raw/irs.zip"
WORK_DIR = "data/raw/_irs_unzipped"
OUT_CSV = "data/processed/irs_hawaii_migration_1990_2023.csv"

#county fips codes to names, 005 (Kalawao) counts as Maui
county_names = {
    "001": "Hawaii County",
    "003": "Honolulu County",
    "005": "Maui County",
    "007": "Kauai County",
    "009": "Maui County",
}

#state codes the IRS uses for "not a real state"
foreign_codes = ["57", "58", "59"]


#helper: turn a cell into a number, IRS uses -1 for suppressed values
def to_number(value):
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if number < 0:
        return None
    return number


#helper: '15.0' -> '15' and '001' -> '1', because xls files store fips codes
#as text in some years and as floats in other years
def clean_code(value):
    text = str(value).strip()
    if text == "":
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return text


#helper: read any sheet of any xls or xlsx as a list of rows of strings
def read_sheet(path, sheet_name_contains):
    rows = []
    if str(path).lower().endswith(".xlsx"):
        book = load_workbook(path, read_only=True, data_only=True)
        sheet_name = book.sheetnames[0]
        for name in book.sheetnames:
            if sheet_name_contains in name.lower():
                sheet_name = name
                break
        for row in book[sheet_name].iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
    else:
        book = xlrd.open_workbook(path, on_demand=True)
        sheet_name = book.sheet_names()[0]
        for name in book.sheet_names():
            if sheet_name_contains in name.lower():
                sheet_name = name
                break
        sheet = book.sheet_by_name(sheet_name)
        for r in range(sheet.nrows):
            rows.append([str(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
    return rows


#function to clean the 1990-1992 text files (C9091hio.txt style)
#these don't have summary rows, so different-state = total - same-state - foreign,
#adding up the destination rows myself
def clean_irs_txt(path, year):
    #the 3 kinds of lines I need to recognize
    header_line = re.compile(r"^(\d{2})\s+(\d{3})\s+(.+?)\s+Total Migration.*?\s+(-?[\d,]+)\s+(-?[\d,]+)\s*$")
    nonmig_line = re.compile(r"^\s*(\d{2})\s+(\d{3})\s+County Non-Migrants\s+(-?[\d,]+)\s+(-?[\d,]+)\s*$")
    dest_line = re.compile(r"^\s+(\d{2})\s+(\d{3})\s+(.+?)\s+([A-Za-z]{2})\s+(-?[\d,]+)\s+[\d.]*\s*(-?[\d,]+)\s+[\d.]*\s*$")

    cleaned = []
    current = None
    for line in Path(path).read_text(errors="replace").splitlines():
        #a county block starts with its Total Migration line
        found = header_line.match(line)
        if found and found.group(1) == "15":
            current = {
                "county_fips": found.group(2),
                "year": year,
                "total_returns": to_number(found.group(4)),
                "total_exemptions": to_number(found.group(5)),
                "same_returns": 0.0, "same_exemptions": 0.0,
                "foreign_returns": 0.0, "foreign_exemptions": 0.0,
            }
            continue
        #a county block ends with its Non-Migrants line
        found = nonmig_line.match(line)
        if found and current is not None and found.group(2) == current["county_fips"]:
            current["nonmig_returns"] = to_number(found.group(3))
            current["nonmig_exemptions"] = to_number(found.group(4))
            cleaned.append(current)
            current = None
            continue
        #in between are the destination rows, add up same-state and foreign
        if current is not None:
            found = dest_line.match(line)
            if found:
                dest_state = found.group(1)
                returns = to_number(found.group(5)) or 0
                exemptions = to_number(found.group(6)) or 0
                if dest_state == "15":
                    current["same_returns"] += returns
                    current["same_exemptions"] += exemptions
                elif dest_state in foreign_codes:
                    current["foreign_returns"] += returns
                    current["foreign_exemptions"] += exemptions

    #now compute different-state as the leftover
    finished = []
    for c in cleaned:
        finished.append(make_row(c))
    return finished


#function to clean the 1992-1994 xls files (C9293Hio.xls style)
#same idea as the text files (no summary rows) just in a spreadsheet:
#the row with destination state '00' is the county total,
#the '63 050' row is non-migrants, '63 011'-'63 014' are region subtotals (skip),
#state 15 rows are same-state and 57/58/59 rows are foreign
def clean_irs_old_xls(path, year):
    counties = {}
    for row in read_sheet(path, "outflow"):
        if len(row) < 8:
            continue
        own_state = clean_code(row[0])
        own_county = clean_code(row[1]).zfill(3)
        if own_state != "15" or own_county not in county_names:
            continue
        dest_state = clean_code(row[2]).zfill(2)
        dest_sub = clean_code(row[3]).zfill(3)
        returns = to_number(row[6])
        exemptions = to_number(row[7])

        if own_county not in counties:
            counties[own_county] = {
                "county_fips": own_county, "year": year,
                "total_returns": None, "total_exemptions": None,
                "same_returns": 0.0, "same_exemptions": 0.0,
                "foreign_returns": 0.0, "foreign_exemptions": 0.0,
                "nonmig_returns": None, "nonmig_exemptions": None,
            }
        current = counties[own_county]

        if dest_state == "00":
            current["total_returns"] = returns
            current["total_exemptions"] = exemptions
        elif dest_state == "63" and dest_sub == "050":
            current["nonmig_returns"] = returns
            current["nonmig_exemptions"] = exemptions
        elif dest_state == "63":
            continue  #region subtotal rows, already inside the total
        elif dest_state == "15":
            current["same_returns"] += returns or 0
            current["same_exemptions"] += exemptions or 0
        elif dest_state in foreign_codes:
            current["foreign_returns"] += returns or 0
            current["foreign_exemptions"] += exemptions or 0

    finished = []
    for c in counties.values():
        finished.append(make_row(c))
    return finished


#helper for the 2 functions above: different-state = total - same - foreign
def make_row(c):
    diff_returns = None
    diff_exemptions = None
    if c["total_returns"] is not None:
        diff_returns = max(c["total_returns"] - c["same_returns"] - c["foreign_returns"], 0)
    if c["total_exemptions"] is not None:
        diff_exemptions = max(c["total_exemptions"] - c["same_exemptions"] - c["foreign_exemptions"], 0)
    return {
        "county_fips": c["county_fips"], "year": c["year"],
        "total_us_foreign_returns": c["total_returns"],
        "total_us_foreign_exemptions": c["total_exemptions"],
        "same_state_returns": c["same_returns"],
        "same_state_exemptions": c["same_exemptions"],
        "diff_state_returns": diff_returns,
        "diff_state_exemptions": diff_exemptions,
        "foreign_returns": c["foreign_returns"],
        "foreign_exemptions": c["foreign_exemptions"],
        "nonmigrant_returns": c["nonmig_returns"],
        "nonmigrant_exemptions": c["nonmig_exemptions"],
    }


#function to clean 1995 onward: these files HAVE summary rows per county,
#coded in the destination-state column:
#  96 = total migration US + foreign
#  97 sub 0 = total US, 97 sub 1 = SAME state, 97 sub 3 = DIFFERENT state
#  98 = foreign
#  and a row whose label says Non-Migrants
#works for both the 1995-2011 single-sheet files and the modern
#County Outflow / County Inflow sheet files (2011 onward)
def clean_irs_xls(path, year, direction):
    if direction == "out":
        sheet_hint = "county outflow"
    else:
        sheet_hint = "county inflow"

    counties = {}
    for row in read_sheet(path, sheet_hint):
        if len(row) < 8:
            continue
        own_state = clean_code(row[0])
        own_county = clean_code(row[1])
        dest_state = clean_code(row[2])
        dest_sub = clean_code(row[3])
        label = str(row[5])

        #in INFLOW files the hawaii county sits in columns 2/3 instead of 0/1
        if direction == "in":
            own_state, own_county, dest_state, dest_sub = dest_state, dest_sub, own_state, own_county
            #except the summary rows keep their codes on the origin side, swap back
            if own_state in ["96", "97", "98"]:
                own_state, own_county, dest_state, dest_sub = dest_state, dest_sub, own_state, own_county

        if own_state != "15" or own_county in ["", "0"]:
            continue  #skip state-level rows
        fips = own_county.zfill(3)
        if fips not in county_names:
            continue

        returns = to_number(row[6])
        exemptions = to_number(row[7])
        if fips not in counties:
            counties[fips] = {}
        current = counties[fips]

        if "non-migrant" in label.lower():
            current["nonmig"] = (returns, exemptions)
        elif dest_state == "96":
            current["total"] = (returns, exemptions)
        elif dest_state == "97" and dest_sub in ["0", "000"]:
            current["total_us"] = (returns, exemptions)
        elif dest_state == "97" and dest_sub in ["1", "001"]:
            current["same"] = (returns, exemptions)
        elif dest_state == "97" and dest_sub in ["3", "003"]:
            current["diff"] = (returns, exemptions)
        elif dest_state == "98":
            current["foreign"] = (returns, exemptions)

    finished = []
    for fips, current in counties.items():
        empty = (None, None)
        finished.append({
            "county_fips": fips, "year": year,
            "total_us_foreign_returns": current.get("total", empty)[0],
            "total_us_foreign_exemptions": current.get("total", empty)[1],
            "same_state_returns": current.get("same", empty)[0],
            "same_state_exemptions": current.get("same", empty)[1],
            "diff_state_returns": current.get("diff", empty)[0],
            "diff_state_exemptions": current.get("diff", empty)[1],
            "foreign_returns": current.get("foreign", empty)[0],
            "foreign_exemptions": current.get("foreign", empty)[1],
            "nonmigrant_returns": current.get("nonmig", empty)[0],
            "nonmigrant_exemptions": current.get("nonmig", empty)[1],
        })
    return finished


#unzip everything (the bundle has zips inside the zip)
work = Path(WORK_DIR)
if not work.exists():
    work.mkdir(parents=True)
    with zipfile.ZipFile(RAW_ZIP) as z:
        z.extractall(work)
    for inner_zip in list(work.rglob("*.zip")):
        with zipfile.ZipFile(inner_zip) as z:
            z.extractall(inner_zip.parent / inner_zip.stem)

#calls
all_rows = []

#part 1: the year folders "Hawaii 1990-1991" ... "Hawaii 2010-2011"
#file names end hio/ohi (outflow) or hii/ihi (inflow), sometimes with an r
#the recompressed bundle can contain the same folder twice, so track seen ones
seen_folders = []
for folder in sorted(work.rglob("Hawaii [0-9]*-[0-9]*")):
    if not folder.is_dir() or folder.name in seen_folders:
        continue
    seen_folders.append(folder.name)
    year = int(folder.name.split()[1].split("-")[0])
    for file in folder.iterdir():
        stem = file.name.lower().split(".")[0].rstrip("r")
        if stem.endswith("hio") or stem.endswith("ohi"):
            direction = "out"
        elif stem.endswith("hii") or stem.endswith("ihi"):
            direction = "in"
        else:
            raise ValueError("can't tell direction of " + file.name)

        if file.name.lower().endswith(".txt"):
            rows = clean_irs_txt(file, year)
        else:
            rows = clean_irs_xls(file, year, direction)
            #if the summary-code parse found nothing this is a 1992-1994
            #old style file, use the fallback
            found_any = False
            for r in rows:
                if r["total_us_foreign_returns"] is not None:
                    found_any = True
            if not found_any:
                rows = clean_irs_old_xls(file, year)

        for r in rows:
            r["direction"] = direction
            all_rows.append(r)

#part 2: the modern flat files 1112hi.xls ... 2223hi.xlsx (both directions inside)
seen_files = []
for file in sorted(work.rglob("*hi.xls*")):
    matched = re.match(r"^(\d{2})(\d{2})hi\.(xls|xlsx)$", file.name.lower())
    if not matched or file.name.lower() in seen_files:
        continue
    seen_files.append(file.name.lower())
    year = 2000 + int(matched.group(1))
    for direction in ["out", "in"]:
        rows = clean_irs_xls(file, year, direction)
        for r in rows:
            r["direction"] = direction
            all_rows.append(r)

#assemble all years into 1 df
irs_df = pd.DataFrame(all_rows)
irs_df["county"] = irs_df["county_fips"].map(county_names)

#add kalawao into maui (they have the same county name now so groupby does it)
value_columns = []
for column in irs_df.columns:
    if column.endswith("returns") or column.endswith("exemptions"):
        value_columns.append(column)
irs_df = irs_df.groupby(["county", "year", "direction"], as_index=False)[value_columns].sum(min_count=1)

#mark the 2 IRS methodology breaks for the era dummies later
irs_df["era"] = "pre2011"
irs_df.loc[irs_df["year"] >= 2011, "era"] = "2011_2021"
irs_df.loc[irs_df["year"] >= 2022, "era"] = "post2022"

irs_df = irs_df.sort_values(["direction", "county", "year"])
irs_df.to_csv(OUT_CSV, index=False)

#lemme seee it
print("wrote", OUT_CSV, "with", len(irs_df), "rows, years", irs_df["year"].min(), "-", irs_df["year"].max())
outflows = irs_df[irs_df["direction"] == "out"].copy()
outflows["out_rate"] = outflows["diff_state_returns"] / (outflows["nonmigrant_returns"] + outflows["total_us_foreign_returns"])
print((outflows.pivot_table(index="year", columns="county", values="out_rate") * 100).round(2).to_string())
