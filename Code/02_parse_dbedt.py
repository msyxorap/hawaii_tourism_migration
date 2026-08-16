#Imports
import re
import zipfile
import unicodedata
import pandas as pd
import xlrd
from openpyxl import load_workbook
from pathlib import Path

# DBEDT Annual Visitor Research Report supporting tables, 1999-2024
# every yearly workbook has a "VISITOR DAYS BY ISLAND" table with 2 columns:
# the report year and the year before. so 1999 and 2003 (the 2 pdf-only
# report years) still get covered by the prior-year column of the 2000 and
# 2004 workbooks.
# when a year shows up in 2 reports the later report wins (it has the
# revised number).
# the island labels changed over the years (BIG ISLAND vs Hawaii Island,
# okina characters on/off) so the labels get normalized first. so much of
# the data was variable since they changed the format over and over and
# over again holy headache (again)

#where the data lives
RAW_ZIP = "data/raw/dbedt.zip"
WORK_DIR = "data/raw/_dbedt_unzipped"
OUT_CSV = "data/processed/dbedt_hawaii_visitor_days_1999_2024.csv"

#island names to county names
island_to_county = {
    "TOTAL STATE": "State", "STATE TOTAL": "State",
    "OAHU": "Honolulu County",
    "MAUI COUNTY": "Maui County",
    "KAUAI": "Kauai County",
    "HAWAII ISLAND": "Hawaii County", "BIG ISLAND": "Hawaii County",
}

#helper: uppercase, strip the okina/accent characters so labels match
def clean_label(value):
    text = unicodedata.normalize("NFKD", str(value))
    letters = []
    for character in text:
        if not unicodedata.combining(character):
            letters.append(character)
    text = "".join(letters).upper()
    for quote in ["'", "\u2018", "\u2019"]:
        text = text.replace(quote, "")
    return re.sub(r"\s+", " ", text).strip()


#helper: read the first 40 rows of every sheet looking for the annual
#VISITOR DAYS BY ISLAND table (not the monthly one, not the MMA one)
def find_visitor_days_rows(path):
    if str(path).lower().endswith(".xlsx"):
        book = load_workbook(path, read_only=True, data_only=True)
        for name in book.sheetnames:
            rows = []
            for row in book[name].iter_rows(max_row=40, max_col=10, values_only=True):
                rows.append(["" if v is None else str(v) for v in row])
            if rows and is_the_right_table(rows):
                return rows
    else:
        book = xlrd.open_workbook(path, on_demand=True)
        for name in book.sheet_names():
            sheet = book.sheet_by_name(name)
            rows = []
            for r in range(min(sheet.nrows, 40)):
                rows.append([str(sheet.cell_value(r, c)) for c in range(min(sheet.ncols, 10))])
            if rows and is_the_right_table(rows):
                return rows
    return None


#helper for the helper: the annual table's title says VISITOR DAYS BY ISLAND
#with no MONTH and no MMA in it
def is_the_right_table(rows):
    title = ""
    for value in rows[0]:
        if value:
            title = title + " " + str(value)
    title = title.upper()
    return "VISITOR DAYS BY ISLAND" in title and "MONTH" not in title and "MMA" not in title


#function to clean one workbook's visitor days table
def clean_dbedt_year(path):
    rows = find_visitor_days_rows(path)
    if rows is None:
        return []

    #find the header row: it has 2 four-digit years in it
    header_index = None
    year_columns = []
    for i in range(len(rows)):
        found_years = []
        for c in range(len(rows[i])):
            text = str(rows[i][c]).strip()
            if re.match(r"^(19|20)\d{2}(\.0)?$", text):
                found_years.append((c, int(float(text))))
        if len(found_years) >= 2:
            header_index = i
            year_columns = found_years[:2]
            break
    if header_index is None:
        return []

    #walk the island rows under the header
    cleaned = []
    for i in range(header_index + 1, min(header_index + 16, len(rows))):
        label = clean_label(rows[i][0]) if rows[i] else ""
        if label.startswith("TABLE") or label.startswith("SOURCE"):
            break
        if label not in island_to_county:
            continue
        for column, year in year_columns:
            try:
                value = float(str(rows[i][column]).replace(",", ""))
            except (TypeError, ValueError):
                continue
            cleaned.append({
                "county": island_to_county[label],
                "year": year,
                "visitor_days": value,
            })
    return cleaned


#unzip
work = Path(WORK_DIR)
if not work.exists():
    work.mkdir(parents=True)
    with zipfile.ZipFile(RAW_ZIP) as z:
        z.extractall(work)

#calls
#every workbook reports 2 years, keep track of which report each number
#came from so the later (revised) report can win
values = {}
seen_reports = []
all_files = sorted(work.rglob("*.xls")) + sorted(work.rglob("*.xlsx"))
for file in all_files:
    matched = re.match(r"^(\d{4})", file.name)
    if not matched:
        continue
    report_year = int(matched.group(1))
    if report_year in seen_reports:
        continue  #the 2020 file is in the zip twice
    rows = clean_dbedt_year(file)
    if len(rows) == 0:
        continue
    seen_reports.append(report_year)
    for r in rows:
        key = (r["county"], r["year"])
        if key not in values or report_year > values[key]["report_year"]:
            values[key] = {"visitor_days": r["visitor_days"], "report_year": report_year}

#assemble all years into 1 df
all_rows = []
for (county, year), v in values.items():
    all_rows.append({
        "county": county, "year": year,
        "visitor_days": v["visitor_days"],
        "source_report_year": v["report_year"],
    })
dbedt_df = pd.DataFrame(all_rows)
dbedt_df = dbedt_df.sort_values(["county", "year"])
Path("data/processed").mkdir(parents=True, exist_ok=True)
dbedt_df.to_csv(OUT_CSV, index=False)

#lemme seee it
print("wrote", OUT_CSV, "with", len(dbedt_df), "rows, years", dbedt_df["year"].min(), "-", dbedt_df["year"].max())
counties_only = dbedt_df[dbedt_df["county"] != "State"]
print((counties_only.pivot_table(index="year", columns="county", values="visitor_days") / 1e6).round(1).to_string())
