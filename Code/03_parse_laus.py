#Imports
import zipfile
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import io

# BLS LAUS county unemployment for Hawaii, 1990-2025
# one laucntyXX.xlsx file per year inside the zip, all the same layout:
# LAUS Code | State FIPS | County FIPS | County Name | Year | Labor Force |
# Employed | Unemployed | Unemployment Rate (%)
# I just keep the State FIPS 15 (Hawaii) rows from each file and stack them.

#where the data lives
RAW_ZIP = "data/raw/laus.zip"
OUT_CSV = "data/processed/laus_hawaii_unemployment_1990_2025.csv"

#county fips to names, same mapping as the IRS script
county_names = {
    "001": "Hawaii County",
    "003": "Honolulu County",
    "005": "Maui County",   #Kalawao counts as Maui here too
    "007": "Kauai County",
    "009": "Maui County",
}

#function to clean one laucnty file
def clean_laus_year(file_bytes):
    book = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    cleaned = []
    for row in sheet.iter_rows(values_only=True):
        if row is None or len(row) < 9:
            continue
        state_fips = str(row[1]).strip() if row[1] is not None else ""
        if state_fips != "15":
            continue
        county_fips = str(row[2]).strip().zfill(3)
        if county_fips not in county_names:
            continue
        cleaned.append({
            "county_fips": county_fips,
            "county": county_names[county_fips],
            "year": int(row[4]),
            "labor_force": row[5],
            "employed": row[6],
            "unemployed": row[7],
            "unemployment_rate": float(row[8]) / 100.0,  #4.5 -> 0.045 like my ACS numbers
        })
    return cleaned

#calls
all_rows = []
with zipfile.ZipFile(RAW_ZIP) as z:
    for name in sorted(z.namelist()):
        if name.lower().endswith(".xlsx") and "laucnty" in name.lower():
            all_rows = all_rows + clean_laus_year(z.read(name))

#assemble all years into 1 df
laus_df = pd.DataFrame(all_rows)

#kalawao is suppressed or tiny, sum it into maui
#(rate can't just be summed so recompute it from the counts)
laus_df = laus_df.groupby(["county", "year"], as_index=False)[["labor_force", "employed", "unemployed"]].sum(min_count=1)
laus_df["unemployment_rate"] = laus_df["unemployed"] / laus_df["labor_force"]

laus_df = laus_df.sort_values(["county", "year"])
Path("data/processed").mkdir(parents=True, exist_ok=True)
laus_df.to_csv(OUT_CSV, index=False)

#lemme seee it
print("wrote", OUT_CSV, "with", len(laus_df), "rows, years", laus_df["year"].min(), "-", laus_df["year"].max())
print((laus_df.pivot_table(index="year", columns="county", values="unemployment_rate") * 100).round(1).to_string())
